# to_csv.py

r'''Input:

 - musicxml .mxl file
 - sections

Output:

 - title.xls
   - Sheets:
     - Source
       - musicxml filename
       - parts
         - id name abbreviation
         - score_instrument (id instrument-name)
         - midi_device (id port)
         - midi_instrument (id midi-channel midi-program volume pan)
         - divisions staves key time
     - P1.Measures
       - number start_spp key time divisions staves
     - P1.Notes
     - P1.Sections
       - id start (measure.offset)

'''

from pathlib import Path
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment

from ..parse_xml import parse
from ..unroll_repeats import unroll_parts
from ..assign_starts import assign_parts
# FIX delete: from ..tie_notes import tie_parts


Centered = Alignment(horizontal="center")

def read_musicxml(music_file):
    parts = parse(music_file)
    new_parts = unroll_parts(parts)
    assign_parts(new_parts)
    # FIX delete: tie_parts(new_parts)
    return new_parts


Note_fraction = {  # the fractional denominator for the various note types
   'whole': 1,
   'half': 2,
   'quarter': 4,
   'eighth': 8,
   '16th': 16,
   '32nd': 32,
   '64th': 64,
   '128th': 128,
   '256th': 256,
}

class row:
    measure_columns = "measure".split()
    measure_defaults = [None] * len(measure_columns)
    note_columns = "id note color start duration type dot voice staff grace " \
                   "actual_notes normal_notes chord_down chord_up".split()
    note_defaults = [None] * len(note_columns)
    note_defaults[6] = False  # dot
    notation_columns = "slur arpeggiate fermata trill".split()
    notation_defaults = [None] + [False] * (len(notation_columns) - 1)
    articulation_columns = "strong_accent accent staccato detached_legato staccatissimo tenuto".split()
    articulation_defaults = [False] * len(articulation_columns)
    columns = measure_columns + note_columns + notation_columns + articulation_columns
    defaults = measure_defaults + note_defaults + notation_defaults + articulation_defaults

    def __init__(self, measure, note):
        self.values = self.defaults.copy()
        self.set_measure_columns(measure)
        self.set_note_columns(note)

    @classmethod
    def header(cls):
        assert len(cls.columns) == len(cls.defaults)
        return cls.columns

    def set_measure_columns(self, measure):
        self.set("measure", measure.number)

    def set_note_columns(self, note):
        for name in self.note_columns:
            if name == 'grace':
                if note.grace is not None:
                    if hasattr(note.grace, "slash"):
                        self.set(name, "slash")
                    else:
                        self.set(name, True)
            elif name == "actual_notes":
                if note.time_modification is not None:
                    self.set(name, note.time_modification.actual_notes)
            elif name == "normal_notes":
                if note.time_modification is not None:
                    self.set(name, note.time_modification.normal_notes)
            elif name == "type":
                self.set(name, Note_fraction[note.type])
            elif name == "duration":
                if hasattr(note, name):
                    self.set(name, getattr(note, name))
            else:
                self.set(name, getattr(note, name))
        if note.notations is not None:
            self.set_notation_columns(note.notations)

    def set_notation_columns(self, notations):
        for name in self.notation_columns:
            if name == "slur":
                if hasattr(notations, name):
                    self.set(name, f"{notations.slur.type}-{notations.slur.number}")
            elif name == "trill":
                self.set(name, hasattr(notations, "ornaments") and 
                               hasattr(notations.ornaments, "trill_mark"))
            else:
                self.set_true(notations, name)
        if hasattr(notations, "articulations"):
            self.set_articulation_columns(notations.articulations)

    def set_articulation_columns(self, articulations):
        for name in self.articulation_columns:
            self.set_true(articulations, name)

    def set(self, name, value):
        self.values[self.columns.index(name)] = value

    def set_true(self, obj, name):
        self.set(name, hasattr(obj, name))

    def set_value(self, obj, name):
        if hasattr(obj, name):
            self.set(name, getattr(obj, name))

def load_source(ws, parts, music_file):
    #ws.column_dimensions["A"].best_fit = True
    name = "Musicxml filename"
    ws.append([name, str(music_file)])
    ws.column_dimensions["A"].width = len(name)
    ws.append([])
    ws["A3"].value = "Parts:"
    ws.merge_cells("D3:E3")
    ws["D3"].value = "Score Instrument"
    ws["D3"].alignment = Centered
    ws.merge_cells("F3:G3")
    ws["F3"].value = "MIDI Device"
    ws["F3"].alignment = Centered
    ws.merge_cells("H3:L3")
    ws["H3"].value = "MIDI Instrument"
    ws["H3"].alignment = Centered
    ws.merge_cells("O3:P3")
    ws["O3"].value = "Key"
    ws["O3"].alignment = Centered
    row = "id name abbreviation " \
          "id".split()
    row.append("instrument name")
    row.extend("id port " \
               "id".split())
    row.append("MIDI channel")
    row.append("MIDI program")
    row.extend("volume pan "
               "divisions staves fifths mode time".split())
    for col in 2, 4, 8, 9:
        ws.column_dimensions["ABCDEFGHIJKLMNOPQRSTUVWXYZ"[col]].width = len(row[col]) + 0.5
    ws.append(row)
    ws.freeze_panes = f"A{ws.max_row + 1}"
    for info, part in parts:
        row = [info.id, info.part_name, info.part_abbreviation]
        si = info.score_instrument
        row.extend([si.id, si.instrument_name])
        md = info.midi_device
        row.extend([md.id, int(md.port)])
        mi = info.midi_instrument
        row.extend([mi.id, int(mi.midi_channel), int(mi.midi_program), float(mi.volume), float(mi.pan)])
        keys = "divisions staves fifths mode time".split()
        values = {}
        for attr in part[0].attributes:
            for key in keys:
                if key == "fifths":
                    if hasattr(attr, "key"):
                        values["fifths"] = int(attr.key.fifths)
                elif key == "mode":
                    if hasattr(attr, "key"):
                        key = attr.key
                        if hasattr(key, "mode"):
                            values["mode"] = key.mode
                        elif key.fifths >= 0:
                            values["mode"] = "major"
                        else:
                            values["mode"] = "minor"
                elif hasattr(attr, key):
                    value = getattr(attr, key)
                    if key == "time":
                        values[key] = f"{value.beats}/{value.beat_type}"
                    else:
                        values[key] = int(value)
        row.extend([values[key] for key in keys])
        ws.append(row)

def load_measure(ws, measure):
    r'''loads one row of: number start_spp key time divisions staves
    '''
    keys = "number start_spp".split()
    attributes = "key time divisions staves".split()
    values = {}
    for attr in measure.attributes:
        for key in attributes:
            if hasattr(attr, key):
                value = getattr(attr, key)
                if key == "key":
                    if hasattr(value, "mode"):
                        values[key] = f"{value.fifths} {value.mode}"
                    else:
                        values[key] = int(value.fifths)
                elif key == "time":
                    values[key] = f"{value.beats}/{value.beat_type}"
                else:
                    values[key] = int(value)
    for key in keys:
        if hasattr(measure, key):
            value = getattr(measure, key)
            try:
                values[key] = int(value)
            except ValueError:
                values[key] = value
    row = []
    for key in reversed(keys + attributes):
        if key in values:
            row.append(values[key])
        elif row:
            row.append(None)
    row.reverse()
    ws.append(row)

def load_note(ws, measure, note):
    values = row(measure, note).values
    if not isinstance(values[4], int):
        values[4] = int(round(values[4]))
    ws.append(values)

def run():
    parser = argparse.ArgumentParser()
    parser.add_argument("music_file", type=Path)

    args = parser.parse_args()
    #print(f"{args=}")

    parts = read_musicxml(args.music_file)

    wb = Workbook()  # create xlsx workbook
    ws = wb.active   # use default sheet for Source
    ws.title = "Source"
    load_source(ws, parts, args.music_file)

    for info, measures in parts:
        part_id = info.id
        measures_ws = wb.create_sheet(f"{part_id}.Measures")
        measures_ws.append("number start_spp key time divisions staves".split())
        measures_ws.freeze_panes = "A2"
        notes_ws = wb.create_sheet(f"{part_id}.Notes")
        header = row.header()
        notes_ws.append(header)
        for col in 1, 3, 7, 15:
            notes_ws.column_dimensions["ABCDEFGHIJKLMNOPQRSTUVWXYZ"[col]].width = 5.5
        for col in 2, 4, 5, 6, 8, 9, 10, 11, 12, 16, 18, 19, 20, 22, 23, 24:
            notes_ws.column_dimensions["ABCDEFGHIJKLMNOPQRSTUVWXYZ"[col]].width = len(header[col]) + 0.5
        for col in 13,:
            notes_ws.column_dimensions["ABCDEFGHIJKLMNOPQRSTUVWXYZ"[col]].width = len(header[col]) + 1.5
        notes_ws.freeze_panes = "A2"
        sections_ws =  wb.create_sheet(f"{part_id}.Sections")
        sections_ws.append("id start".split())
        sections_ws.freeze_panes = "A2"
        for m in measures:
            load_measure(measures_ws, m)
            for n in m.sorted_notes:
                load_note(notes_ws, m, n)
    wb.save(f"{args.music_file.stem}.xlsx")



if __name__ == "__main__":
    run()
